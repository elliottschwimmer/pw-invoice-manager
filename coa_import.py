"""Imports the City's chart-of-accounts spreadsheet into the app:
  * The per-segment sheets (FUND, DEPARTMENT, DIVISION, SUB-DIVISION,
    PROGRAM, ACTIVITY) -> SegmentCode, so each segment's code can be
    resolved to its name.
  * "Active Accounts" -> ActiveAccount, one row per valid 8-segment GL
    account string, with a search_text field combining every segment's
    name (fund, department, division, ..., object description) so the
    account-search typeahead matches on a fund/division/program NAME
    ("Parking Services") just as well as a code or object description
    ("612990", "contractual").

Run standalone:
    python3 coa_import.py "/path/to/COA.xlsx"

Re-running replaces both tables entirely, so re-importing an updated
spreadsheet just refreshes the data.
"""
import sys

import openpyxl

from models import ActiveAccount, SegmentCode, db

# Sheet name -> (segment key, is the account's 8-segment order position)
_SEGMENT_SHEETS = {
    "FUND": "fund",
    "DEPARTMENT": "dept",
    "DIVISION": "div",
    "SUB-DIVISION": "subdiv",
    "PROGRAM": "program",
    "ACTIVITY": "activity",
}

# Position of each segment within the hyphen-split account string, matching
# the City's 8-segment layout: Fund-Dept-Div-Subdiv-Program-Future-Activity-Object.
_SEGMENT_POSITIONS = ["fund", "dept", "div", "subdiv", "program", None, "activity", None]

# The COA spreadsheet's names are heavily abbreviated (e.g. "PKG SERVICES",
# "TRANS", "MGMT"). Staff will naturally search in plain English ("parking",
# "transportation"), so expand common abbreviations into the search index —
# add both forms rather than replacing, since some rows already spell things
# out. Extend this list as real searches turn up more gaps.
_ABBREVIATIONS = {
    "PKG": "PARKING", "SVCS": "SERVICES", "SVC": "SERVICE",
    "MGMT": "MANAGEMENT", "ADMIN": "ADMINISTRATION", "MAINT": "MAINTENANCE",
    "ENG": "ENGINEERING", "DEPT": "DEPARTMENT", "PROF": "PROFESSIONAL",
    "MISC": "MISCELLANEOUS", "EQUIP": "EQUIPMENT", "SUPP": "SUPPLIES",
    "TRANS": "TRANSPORTATION", "COMM": "COMMUNITY", "CTR": "CENTER",
    "BLDG": "BUILDING", "MGR": "MANAGER", "DEV": "DEVELOPMENT",
    "FAC": "FACILITIES", "REC": "RECREATION", "PW": "PUBLIC WORKS",
    "PD": "POLICE", "FD": "FIRE",
}


def _expand_abbreviations(text: str) -> str:
    if not text:
        return text
    extra = []
    for word in text.replace("-", " ").split():
        expansion = _ABBREVIATIONS.get(word.upper())
        if expansion:
            extra.append(expansion)
    return text + (" " + " ".join(extra) if extra else "")


def import_segment_codes(path: str) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    SegmentCode.query.delete()

    count = 0
    for sheet_name, segment_key in _SEGMENT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            code = str(row[0]).strip()
            description = (row[1] or "").strip() if len(row) > 1 else ""
            if not code:
                continue
            db.session.add(SegmentCode(segment=segment_key, code=code, description=description))
            count += 1

    db.session.commit()
    return count


def import_active_accounts(path: str) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Active Accounts"]

    # Build an in-memory lookup: (segment, code) -> description, so each
    # account row can be decoded without a query per segment per row.
    lookup = {(sc.segment, sc.code): sc.description for sc in SegmentCode.query.all()}

    ActiveAccount.query.delete()

    count = 0
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Record Number, Account, Tp, S, Description, Short Desc
        if not row or not row[1]:
            continue
        account = str(row[1]).strip().rstrip("-")
        if not account or account in seen:
            continue
        seen.add(account)

        object_description = (row[4] or "").strip() if len(row) > 4 else ""

        parts = account.split("-")
        segment_names = []
        for i, seg_key in enumerate(_SEGMENT_POSITIONS):
            if seg_key is None or i >= len(parts):
                continue
            name = lookup.get((seg_key, parts[i]))
            if name:
                segment_names.append(name)

        search_text = " | ".join(segment_names + ([object_description] if object_description else []))
        search_text = _expand_abbreviations(search_text)

        db.session.add(
            ActiveAccount(
                account=account,
                description=object_description,
                short_description=(row[5] or "").strip() if len(row) > 5 else "",
                search_text=search_text,
            )
        )
        count += 1
        if count % 2000 == 0:
            db.session.flush()

    db.session.commit()
    return count


def import_all(path: str):
    seg_count = import_segment_codes(path)
    acct_count = import_active_accounts(path)
    return seg_count, acct_count


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 coa_import.py /path/to/COA.xlsx")
        sys.exit(1)

    from app import create_app
    app = create_app()
    with app.app_context():
        seg_count, acct_count = import_all(sys.argv[1])
        print(f"Imported {seg_count} segment codes and {acct_count} active accounts")
